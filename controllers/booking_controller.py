import os
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from db import get_conn, EXPORTS_DIR
from auth import login_required, manager_required
from models import Booking, Session, Payment, Package, Employee

booking_bp = Blueprint('booking', __name__)

@booking_bp.route("/quick_session", methods=["GET", "POST"])
@login_required
def quick_session():
    error = None
    if request.method == "POST":
        booking_id = request.form.get("booking_id")
        if session.get('employee_role') == 'manager':
            employee_id = request.form.get('employee_id') or session.get('employee_id')
        else:
            employee_id = session.get('employee_id')
        if not booking_id or not employee_id:
            error = "يرجى إدخال رقم العملية ورقم الموظف"
        else:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT customer_id,total_sessions,sessions_done FROM bookings WHERE id=?", (booking_id,))
            b = cur.fetchone()
            if not b:
                error = "لا يوجد عملية بهذا الرقم"
            else:
                customer_id, total, done = b
                if done >= total:
                    error = "هذا الحجز مكتمل"
                else:
                    sn = done + 1
                    today = datetime.now().strftime("%Y-%m-%d")
                    note_val = request.form.get('note')
                    Session.create(booking_id, sn, today, employee_id, 0, note_val)
                    cur.execute("UPDATE bookings SET sessions_done = sessions_done + 1 WHERE id = ?", (booking_id,))
                    conn.commit()
                    conn.close()
                    return redirect(url_for("customer_detail", customer_id=customer_id))
            if conn:
                conn.close()
    return render_template("quick_session.html", error=error)

@booking_bp.route("/bookings/<int:booking_id>/update_next_session", methods=["POST"])
@login_required
def update_next_session(booking_id):
    next_date = request.form.get("next_session_date")
    Booking.update_next_session_date(booking_id, next_date)
    flash("تم تحديث موعد الجلسة القادمة بنجاح", "success")
    return redirect(request.referrer or url_for("index"))

@booking_bp.route("/bookings/<int:booking_id>/add_session", methods=["POST"])
@login_required
def add_session(booking_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT total_sessions,sessions_done,pulses_total,pulses_used FROM bookings WHERE id=?", (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return redirect(url_for("index"))
    total, done, pulses_total, pulses_used = row
    today = datetime.now().strftime("%Y-%m-%d")
    
    if session.get('employee_role') == 'manager':
        employee_id = request.form.get('employee_id') or session.get('employee_id')
    else:
        employee_id = session.get('employee_id')
    try:
        employee_id = int(employee_id)
    except Exception:
        employee_id = session.get('employee_id')
    
    if pulses_total and pulses_total > 0:
        try:
            use = int(request.form.get('pulses_used') or 0)
        except Exception:
            use = 0
        if use <= 0 or pulses_used + use > pulses_total:
            conn.close()
            return ("عدد النبضات غير كافٍ", 400)
        note_val = request.form.get('note')
        Session.create(booking_id, done + 1, today, employee_id, use, note_val)
        cur.execute("UPDATE bookings SET pulses_used = pulses_used + ? WHERE id = ?", (use, booking_id))
    else:
        if done < total:
            session_number = done + 1
            note_val = request.form.get('note')
            Session.create(booking_id, session_number, today, employee_id, 0, note_val)
            cur.execute("UPDATE bookings SET sessions_done = sessions_done + 1 WHERE id = ?", (booking_id,))
        else:
            conn.close()
            return redirect(request.referrer or url_for("index"))
    
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for("index"))

@booking_bp.route("/bookings/<int:booking_id>/pay", methods=["POST"])
@login_required
def add_payment(booking_id):
    method = (request.form.get('method') or '').strip().lower()
    amount_str = request.form.get('amount')
    allowed = {"cash", "wallet", "instapay"}
    try:
        amount = int(amount_str or '0')
    except Exception:
        amount = 0
    
    if method not in allowed or amount <= 0:
        flash("بيانات الدفع غير صحيحة", "danger")
        return redirect(request.referrer or url_for("index"))
    
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''SELECT b.price_override, p.price, (SELECT COALESCE(SUM(pay.amount), 0) FROM payments pay WHERE pay.booking_id = b.id) as total_paid FROM bookings b JOIN packages p ON p.id = b.package_id WHERE b.id = ?''', (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        flash("الحجز غير موجود", "danger")
        return redirect(url_for("index"))
    price_override, package_price, total_paid = row
    actual_price = price_override if price_override is not None else package_price
    remaining = actual_price - total_paid
    
    if amount > remaining:
        conn.close()
        flash(f"لا يمكن دفع مبلغ أكبر من المتبقي ({remaining} جنيه)", "danger")
        return redirect(request.referrer or url_for("index"))
    
    today = datetime.now().strftime("%Y-%m-%d")
    Payment.create(booking_id, amount, method, today, session.get('employee_id'))
    flash("تم تسجيل عملية الدفع بنجاح", "success")
    return redirect(request.referrer or url_for("index"))

@booking_bp.route("/sessions/<int:sid>/delete", methods=["POST"])
@manager_required
def delete_session(sid):
    booking_id = Session.delete_and_update_booking(sid)
    return redirect(request.referrer or url_for("index"))

@booking_bp.route("/sessions/<int:sid>/update_note", methods=["POST"])
@manager_required
def update_session_note(sid):
    note = request.form.get('note') or ''
    Session.update_note(sid, note)
    return redirect(request.referrer or url_for("index"))

@booking_bp.route("/bookings/<int:bid>/delete", methods=["POST"])
@manager_required
def delete_booking(bid):
    customer_id = Booking.delete_with_relations(bid)
    if customer_id:
        return redirect(url_for("customer_detail", customer_id=customer_id))
    else:
        return redirect(request.referrer or url_for("index"))

@booking_bp.route('/invoice/<int:booking_id>')
@login_required
def invoice(booking_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''SELECT b.id,b.customer_id,b.package_id,b.total_sessions,b.sessions_done,b.start_date,b.employee_id,p.name,p.price,c.name,c.phone,c.note FROM bookings b JOIN packages p ON p.id=b.package_id JOIN customers c ON c.id=b.customer_id WHERE b.id=?''', (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return redirect(url_for('index'))
    cur.execute('SELECT id,session_number,date,employee_id FROM sessions WHERE booking_id=? ORDER BY session_number', (booking_id,))
    sessions = cur.fetchall()
    emp_name = None
    if row[6]:
        cur.execute('SELECT name FROM employees WHERE id=?', (row[6],))
        re = cur.fetchone()
        emp_name = re[0] if re else None
    cur.execute('SELECT e.name, COUNT(s.id) FROM sessions s LEFT JOIN employees e ON e.id=s.employee_id WHERE s.booking_id=? GROUP BY e.name', (booking_id,))
    sessions_by_emp = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name', (booking_id,))
    payments_by_emp = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, p.method, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name, p.method', (booking_id,))
    method_rows = cur.fetchall()
    emp_methods = {}
    for n, m, s in method_rows:
        if not n:
            continue
        d = emp_methods.setdefault(n, {'cash': 0, 'wallet': 0, 'instapay': 0})
        if m in d:
            d[m] = s
    emp_detail_list = []
    names = set(list(sessions_by_emp.keys()) + list(payments_by_emp.keys()))
    for n in names:
        ms = emp_methods.get(n, {})
        emp_detail_list.append({
            'name': n,
            'sessions_count': sessions_by_emp.get(n, 0),
            'paid_total': payments_by_emp.get(n, 0),
            'cash': ms.get('cash', 0),
            'wallet': ms.get('wallet', 0),
            'instapay': ms.get('instapay', 0),
        })
    cur.execute('SELECT id,amount,method,date,employee_id FROM payments WHERE booking_id=? ORDER BY id', (booking_id,))
    payments = cur.fetchall()
    conn.close()
    return render_template('invoice.html', booking=row, sessions=sessions, emp_name=emp_name, emp_detail_list=emp_detail_list, payments=payments)

@booking_bp.route('/export_booking_invoice/<int:booking_id>')
@login_required
def export_booking_invoice(booking_id):
    import openpyxl
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''SELECT b.id,b.customer_id,b.package_id,b.total_sessions,b.sessions_done,b.start_date,b.employee_id,p.name,p.price,c.name,c.phone FROM bookings b JOIN packages p ON p.id=b.package_id JOIN customers c ON c.id=b.customer_id WHERE b.id=?''', (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return redirect(url_for('index'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'invoice'
    ws.append(['booking_id','customer_id','customer_name','phone','package','price','total_sessions','sessions_done','start_date','employee'])
    emp_name = None
    if row[6]:
        cur.execute('SELECT name FROM employees WHERE id=?', (row[6],))
        re = cur.fetchone()
        emp_name = re[0] if re else None
    ws.append([row[0], row[1], row[9], row[10], row[7], row[8], row[3], row[4], row[5], emp_name])
    ws2 = wb.create_sheet('sessions')
    ws2.append(['id','session_number','date','employee'])
    cur.execute('SELECT id,session_number,date,employee_id FROM sessions WHERE booking_id=? ORDER BY session_number', (booking_id,))
    for s in cur.fetchall():
        en = None
        if s[3]:
            cur.execute('SELECT name FROM employees WHERE id=?', (s[3],))
            re = cur.fetchone()
            en = re[0] if re else None
        ws2.append([s[0], s[1], s[2], en])
    ws3 = wb.create_sheet('payments')
    ws3.append(['id','amount','method','date','employee'])
    cur.execute('SELECT id,amount,method,date,employee_id FROM payments WHERE booking_id=? ORDER BY id', (booking_id,))
    for p in cur.fetchall():
        en = None
        if p[4]:
            cur.execute('SELECT name FROM employees WHERE id=?', (p[4],))
            re = cur.fetchone()
            en = re[0] if re else None
        ws3.append([p[0], p[1], p[2], p[3], en])
    ws4 = wb.create_sheet('employees')
    ws4.append(['employee','sessions_count','paid_total','cash','wallet','instapay'])
    cur.execute('SELECT e.name, COUNT(s.id) FROM sessions s LEFT JOIN employees e ON e.id=s.employee_id WHERE s.booking_id=? GROUP BY e.name', (booking_id,))
    s_map = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name', (booking_id,))
    pay_map = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, p.method, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name, p.method', (booking_id,))
    rows = cur.fetchall()
    m_map = {}
    for n, m, v in rows:
        if not n:
            continue
        d = m_map.setdefault(n, {'cash': 0, 'wallet': 0, 'instapay': 0})
        if m in d:
            d[m] = v
    names = set(list(s_map.keys()) + list(pay_map.keys()) + list(m_map.keys()))
    for n in names:
        d = m_map.get(n, {})
        ws4.append([n, s_map.get(n, 0), pay_map.get(n, 0), d.get('cash', 0), d.get('wallet', 0), d.get('instapay', 0)])
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f'invoice_{booking_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

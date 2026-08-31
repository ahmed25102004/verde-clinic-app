import os
import json
from datetime import datetime, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from db import get_conn, EXPORTS_DIR
from auth import login_required, manager_required
from models import Expense

report_bp = Blueprint('report', __name__)

@report_bp.route('/report')
@manager_required
def report_daily():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date_single = request.args.get('date')
    
    if date_single:
        start_date = date_single
        end_date = date_single
    
    if not start_date:
        start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date:
        end_date = start_date

    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute('SELECT id,name FROM employees ORDER BY name')
    emps = cur.fetchall()
    
    expenses_map = Expense.get_expenses_map_by_date_range(start_date, end_date)
    
    rows = []
    total_cash = 0
    total_wallet = 0
    total_instapay = 0
    total_expenses_total = 0
    
    for emp in emps:
        eid = emp[0]
        cur.execute('SELECT method, COALESCE(SUM(amount),0) FROM payments WHERE employee_id=? AND date BETWEEN ? AND ? GROUP BY method', (eid, start_date, end_date))
        pms = cur.fetchall()
        pm = {'cash': 0, 'wallet': 0, 'instapay': 0}
        for m in pms:
            if m[0] in pm:
                pm[m[0]] = m[1]
        
        emp_expenses = expenses_map.get(eid, 0)
        total_for_manager = (pm['cash'] + pm['wallet'] + pm['instapay']) - emp_expenses
        
        total_cash += pm['cash']
        total_wallet += pm['wallet']
        total_instapay += pm['instapay']
        total_expenses_total += emp_expenses
        
        cur.execute('SELECT COUNT(id) FROM sessions WHERE employee_id=? AND date BETWEEN ? AND ?', (eid, start_date, end_date))
        sessions_count = cur.fetchone()[0]
        
        cur.execute('''SELECT COUNT(b.id), COALESCE(SUM(p.price),0) FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.employee_id=? AND b.start_date BETWEEN ? AND ?''', (eid, start_date, end_date))
        bc, bv = cur.fetchone()
        
        rows.append({
            'employee': emp[1], 'employee_id': eid, 'cash': pm['cash'], 'wallet': pm['wallet'],
            'instapay': pm['instapay'], 'total_payments': pm['cash'] + pm['wallet'] + pm['instapay'],
            'expenses': emp_expenses, 'total_for_manager': total_for_manager, 'sessions_count': sessions_count,
            'bookings_count': bc or 0, 'bookings_value': bv or 0
        })
    
    grand_total = total_cash + total_wallet + total_instapay
    net_total = grand_total - total_expenses_total
    
    cur.execute('SELECT COALESCE(SUM(amount),0) FROM payments WHERE date BETWEEN ? AND ?', (start_date, end_date))
    payments_total = cur.fetchone()[0]
    
    cur.execute('SELECT method, COALESCE(SUM(amount),0) FROM payments WHERE date BETWEEN ? AND ? GROUP BY method', (start_date, end_date))
    payments_by_method = cur.fetchall()
    
    cur.execute('SELECT COUNT(*) FROM payments WHERE date BETWEEN ? AND ?', (start_date, end_date))
    payments_count = cur.fetchone()[0]
    
    cur.execute('SELECT COUNT(DISTINCT b.customer_id) FROM payments p JOIN bookings b ON b.id=p.booking_id WHERE p.date BETWEEN ? AND ?', (start_date, end_date))
    unique_payers = cur.fetchone()[0]
    
    cur.execute('SELECT COALESCE(SUM(p.price),0), COUNT(b.id) FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.start_date BETWEEN ? AND ?', (start_date, end_date))
    bv_row = cur.fetchone()
    bookings_value_today = bv_row[0] if bv_row else 0
    bookings_count_today = bv_row[1] if bv_row else 0
    
    cur.execute('SELECT COUNT(*) FROM sessions WHERE date BETWEEN ? AND ?', (start_date, end_date))
    sessions_count_today = cur.fetchone()[0]
    
    cur.execute('''SELECT COALESCE(SUM(amount), 0) FROM payments WHERE booking_id IN (SELECT id FROM bookings WHERE start_date BETWEEN ? AND ?)''', (start_date, end_date))
    bookings_paid_sum = cur.fetchone()[0]
    bookings_remaining = bookings_value_today - bookings_paid_sum

    conn.close()
    
    summary = {
        'start_date': start_date, 'end_date': end_date, 'total_cash': total_cash, 'total_wallet': total_wallet,
        'total_instapay': total_instapay, 'total_expenses': total_expenses_total, 'grand_total': grand_total,
        'net_total': net_total, 'payments_total': payments_total or 0, 'payments_by_method': payments_by_method or [],
        'payments_count': payments_count or 0, 'unique_payers': unique_payers or 0, 'bookings_value_today': bookings_value_today or 0,
        'bookings_count_today': bookings_count_today or 0, 'sessions_count_today': sessions_count_today or 0,
        'bookings_paid_sum': bookings_paid_sum, 'bookings_remaining': bookings_remaining
    }
    return render_template('report.html', rows=rows, summary=summary, start_date=start_date, end_date=end_date)

@report_bp.route('/dashboard')
@manager_required
def dashboard():
    conn = get_conn()
    cur = conn.cursor()
    
    today_str = datetime.now().strftime('%Y-%m-%d')
    month_start = datetime.now().strftime('%Y-%m-01')
    tomorrow_str = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

    # KPI Metrics
    cur.execute('SELECT COALESCE(SUM(amount), 0) FROM payments WHERE date = ?', (today_str,))
    today_revenue = cur.fetchone()[0]

    cur.execute('SELECT COALESCE(SUM(amount), 0) FROM payments WHERE date >= ?', (month_start,))
    month_revenue = cur.fetchone()[0]

    cur.execute('SELECT COUNT(*) FROM sessions WHERE date = ?', (today_str,))
    today_sessions_count = cur.fetchone()[0]

    cur.execute('SELECT COUNT(*) FROM customers')
    total_customers_count = cur.fetchone()[0]

    cur.execute('SELECT COUNT(*) FROM bookings WHERE sessions_done < total_sessions')
    active_bookings_count = cur.fetchone()[0]

    cur.execute('''SELECT COALESCE(SUM(COALESCE(b.price_override, pkg.price) - (SELECT COALESCE(SUM(p.amount), 0) FROM payments p WHERE p.booking_id = b.id)), 0) FROM bookings b JOIN packages pkg ON b.package_id = pkg.id WHERE (COALESCE(b.price_override, pkg.price) - (SELECT COALESCE(SUM(p.amount), 0) FROM payments p WHERE p.booking_id = b.id)) > 0''')
    total_uncollected_debt = cur.fetchone()[0]

    # Today payments breakdown by method
    cur.execute('SELECT method, COALESCE(SUM(amount), 0) FROM payments WHERE date = ? GROUP BY method', (today_str,))
    today_methods_raw = dict(cur.fetchall())
    today_payment_methods = {
        'cash': today_methods_raw.get('cash', 0),
        'wallet': today_methods_raw.get('wallet', 0),
        'instapay': today_methods_raw.get('instapay', 0)
    }

    # Revenue last 7 days
    last_7_days = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        cur.execute('SELECT COALESCE(SUM(amount), 0) FROM payments WHERE date=?', (d,))
        last_7_days.append({'date': d, 'amount': cur.fetchone()[0]})
    
    # Employee Stats
    cur.execute('''SELECT e.name, (SELECT COUNT(*) FROM sessions WHERE employee_id = e.id) as sessions, (SELECT COALESCE(SUM(amount), 0) FROM payments WHERE employee_id = e.id) as revenue FROM employees e ORDER BY revenue DESC''')
    employee_stats = cur.fetchall()
    
    # Top Packages
    cur.execute('''SELECT p.name, COUNT(b.id) as bookings_count, p.price FROM packages p LEFT JOIN bookings b ON p.id = b.package_id GROUP BY p.id ORDER BY bookings_count DESC LIMIT 5''')
    top_packages = cur.fetchall()
    
    # Top Customers
    cur.execute('''SELECT c.name, c.phone, COALESCE(SUM(p.amount), 0) as total_paid FROM customers c JOIN bookings b ON c.id = b.customer_id JOIN payments p ON b.id = p.booking_id GROUP BY c.id ORDER BY total_paid DESC LIMIT 5''')
    top_customers = cur.fetchall()

    # Upcoming appointments today & tomorrow with WhatsApp payload
    cur.execute('''SELECT c.name, c.phone, b.next_session_date, p.name, b.id, c.id FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages p ON b.package_id = p.id WHERE b.next_session_date = ? OR b.next_session_date = ? ORDER BY b.next_session_date ASC''', (today_str, tomorrow_str))
    appointments_raw = cur.fetchall()
    upcoming_appointments_list = []
    for appt in appointments_raw:
        c_name, c_phone, n_date, p_name, b_id, c_id = appt
        is_today = (n_date == today_str)
        day_label = "اليوم" if is_today else "غداً"
        formatted_phone = (c_phone or '').strip()
        if formatted_phone.startswith('0'):
            formatted_phone = '2' + formatted_phone
        elif len(formatted_phone) == 10 and formatted_phone.startswith('1'):
            formatted_phone = '20' + formatted_phone
            
        wa_msg = f"مرحباً {c_name}، نذكرك بموعد جلستك القادمة ({p_name}) {day_label} في مركزنا. ننتظرك بكل حب."
        upcoming_appointments_list.append({
            'customer_name': c_name,
            'customer_id': c_id,
            'booking_id': b_id,
            'phone': c_phone,
            'formatted_phone': formatted_phone,
            'package_name': p_name,
            'date': n_date,
            'day_label': day_label,
            'is_today': is_today,
            'wa_msg': wa_msg
        })

    # Insights
    insights = []
    cur.execute('''SELECT c.name, c.id, (COALESCE(b.price_override, pkg.price) - (SELECT COALESCE(SUM(amount), 0) FROM payments WHERE booking_id = b.id)) as remaining FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE remaining > 500 ORDER BY remaining DESC LIMIT 3''')
    debtors = cur.fetchall()
    for d in debtors:
        insights.append({'type': 'warning', 'icon': 'bi-exclamation-triangle-fill', 'text': f"العميلة {d[0]} عليها مبلغ متبقي قدره ({d[2]} ج.م). يُنصح بالمتابعة لتسديد المتبقي."})

    cur.execute('''SELECT c.name, pkg.name FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE b.sessions_done >= b.total_sessions AND NOT EXISTS (SELECT 1 FROM bookings b2 WHERE b2.customer_id = b.customer_id AND b2.id > b.id) LIMIT 3''')
    finished = cur.fetchall()
    for f in finished:
        insights.append({'type': 'info', 'icon': 'bi-check-circle-fill', 'text': f"العميلة {f[0]} أتمت جميع جلسات {f[1]}. فرصة ممتازة لعرض باكدج جديد."})

    if today_sessions_count == 0:
        insights.append({'type': 'danger', 'icon': 'bi-alarm-fill', 'text': "لم يتم تسجيل أي جلسات تنفيذية اليوم حتى الآن."})

    conn.close()
    return render_template('dashboard.html',
                         today_revenue=today_revenue,
                         month_revenue=month_revenue,
                         today_sessions_count=today_sessions_count,
                         total_customers_count=total_customers_count,
                         active_bookings_count=active_bookings_count,
                         total_uncollected_debt=total_uncollected_debt,
                         today_payment_methods=today_payment_methods,
                         last_7_days=last_7_days,
                         employee_stats=employee_stats,
                         top_packages=top_packages,
                         top_customers=top_customers,
                         upcoming_appointments=upcoming_appointments_list,
                         insights=insights)

@report_bp.route("/customers")
@manager_required
def customers_list():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""SELECT c.id, c.name, c.phone, c.note, c.created_at, COUNT(b.id) AS bookings_count FROM customers c LEFT JOIN bookings b ON b.customer_id = c.id GROUP BY c.id ORDER BY c.id DESC""")
    customers = cur.fetchall()
    conn.close()
    return render_template("customers.html", customers=customers, rows=customers)

@report_bp.route("/packages", methods=["GET", "POST"])
@manager_required
def packages_admin():
    conn = get_conn()
    cur = conn.cursor()
    error = None
    if request.method == "POST":
        category = request.form.get("category")
        name = request.form.get("name")
        sessions_count = request.form.get("sessions_count")
        price = request.form.get("price")
        bonus = request.form.get("bonus")
        allowed = {"cosmetic_sessions", "cosmetic_packages", "laser_sessions", "laser_packages", "pulse_packages"}
        try:
            sc = int(sessions_count or 0)
            pr = int(price or 0)
            if category not in allowed or not name or sc <= 0 or pr <= 0:
                error = "بيانات غير صحيحة"
            else:
                cur.execute("INSERT INTO packages(category,name,sessions_count,price,bonus) VALUES(?,?,?,?,?)", (category, name, sc, pr, bonus))
                conn.commit()
        except ValueError:
            error = "القيم الرقمية غير صحيحة"
    cur.execute("SELECT id,category,name,sessions_count,price FROM packages WHERE (is_active = 1 OR is_active IS NULL) ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return render_template("packages.html", rows=rows, error=error)

@report_bp.route("/packages/delete", methods=["GET", "POST"])
@manager_required
def packages_delete():
    if request.method == "GET":
        return redirect(url_for("report.packages_admin"))
        
    pid = request.form.get("id")
    error = None
    try:
        p = int(pid)
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("UPDATE packages SET is_active = 0 WHERE id = ?", (p,))
        conn.commit()
        conn.close()
        flash("تم حذف الباكدج بنجاح من قائمة الإتاحة، مع الاحتفاظ ببيانات العملاء المشتركين بها سابقاً", "success")
        return redirect(url_for("report.packages_admin"))
    except (TypeError, ValueError):
        error = "معرف الباكدج غير صالح"
    except Exception as e:
        error = f"خطأ عند الحذف: {str(e)}"
        
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,category,name,sessions_count,price FROM packages WHERE (is_active = 1 OR is_active IS NULL) ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return render_template("packages.html", rows=rows, error=error)

@report_bp.route("/export")
@manager_required
def export_all():
    import openpyxl
    wb = openpyxl.Workbook()
    ws_emp = wb.active
    ws_emp.title = "employees"
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name FROM employees ORDER BY id")
    ws_emp.append(["id", "name"])
    for r in cur.fetchall():
        ws_emp.append(list(r))
    ws_cus = wb.create_sheet("customers")
    ws_cus.append(["id", "name", "phone", "note", "created_at"])
    cur.execute("SELECT id,name,phone,note,created_at FROM customers ORDER BY id")
    for r in cur.fetchall():
        ws_cus.append(list(r))
    ws_pkg = wb.create_sheet("packages")
    ws_pkg.append(["id", "category", "name", "sessions_count", "price", "bonus"])
    cur.execute("SELECT id,category,name,sessions_count,price,bonus FROM packages ORDER BY id")
    for r in cur.fetchall():
        ws_pkg.append(list(r))
    ws_b = wb.create_sheet("bookings")
    ws_b.append(["id", "customer_id", "package_id", "total_sessions", "sessions_done", "start_date", "employee_id", "pulses_total", "pulses_used"])
    cur.execute("SELECT id,customer_id,package_id,total_sessions,sessions_done,start_date,employee_id,pulses_total,pulses_used FROM bookings ORDER BY id")
    for r in cur.fetchall():
        ws_b.append(list(r))
    ws_s = wb.create_sheet("sessions")
    ws_s.append(["id", "booking_id", "session_number", "date", "employee_id", "pulses_used"])
    cur.execute("SELECT id,booking_id,session_number,date,employee_id,pulses_used FROM sessions ORDER BY id")
    for r in cur.fetchall():
        ws_s.append(list(r))
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"pos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@report_bp.route("/report_export")
@manager_required
def report_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    start_date = request.args.get("start_date") or datetime.now().strftime("%Y-%m-%d")
    end_date = request.args.get("end_date") or start_date
    
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""SELECT c.name, pkg.category, pkg.name, p.amount, p.date, p.method FROM payments p JOIN bookings b ON p.booking_id = b.id JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE p.date BETWEEN ? AND ? ORDER BY p.date ASC, p.id ASC""", (start_date, end_date))
    payments = cur.fetchall()
    
    cur.execute("""SELECT method, COALESCE(SUM(amount), 0) FROM payments WHERE date BETWEEN ? AND ? GROUP BY method""", (start_date, end_date))
    method_totals = {row[0]: row[1] for row in cur.fetchall()}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.rightToLeft = True
    
    headers = ["اسم العميل", "الفئة (Area)", "الباكدج", "المبلغ المدفوع", "التاريخ", "طريقة الدفع"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for p in payments:
        method_ar = "نقدي" if p[5] == 'cash' else "محفظة" if p[5] == 'wallet' else "انستا باي" if p[5] == 'instapay' else p[5]
        ws.append([p[0], p[1], p[2], p[3], p[4], method_ar])
    
    ws.append([])
    ws.append(["إجمالي النقدي", method_totals.get('cash', 0), "جنيه"])
    ws.append(["إجمالي المحفظة", method_totals.get('wallet', 0), "جنيه"])
    ws.append(["إجمالي انستا باي", method_totals.get('instapay', 0), "جنيه"])
    ws.append(["الإجمالي الكلي", sum(method_totals.values()), "جنيه"])
    
    last_row = ws.max_row
    for i in range(last_row - 3, last_row + 1):
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2).font = Font(bold=True)
    
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"report_{start_date}_to_{end_date}.xlsx"
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@report_bp.route("/export_daily_customers")
@manager_required
def export_daily_customers():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("""SELECT c.name as customer_name, pkg.name as package_name, p.amount as paid_amount, p.method as payment_method, p.date as transaction_date, b.id as booking_id, COALESCE(b.price_override, pkg.price) as total_price FROM payments p JOIN bookings b ON p.booking_id = b.id JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE p.date = ?""", (date,))
    payments_today = cur.fetchall()
    
    cur.execute("""SELECT c.name as customer_name, pkg.name as package_name, 0 as paid_amount, 'لم يدفع' as payment_method, b.start_date as transaction_date, b.id as booking_id, COALESCE(b.price_override, pkg.price) as total_price FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE b.start_date = ? AND b.id NOT IN (SELECT booking_id FROM payments WHERE date = ?)""", (date, date))
    bookings_without_payment_today = cur.fetchall()

    cur.execute("""SELECT DISTINCT c.name as customer_name, pkg.name as package_name, 0 as paid_amount, 'لم يدفع' as payment_method, b.start_date as transaction_date, b.id as booking_id, COALESCE(b.price_override, pkg.price) as total_price FROM sessions s JOIN bookings b ON s.booking_id = b.id JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE s.date = ? AND b.id NOT IN (SELECT booking_id FROM payments WHERE date = ?) AND b.start_date != ?""", (date, date, date))
    sessions_only_today = cur.fetchall()
    
    all_records = list(payments_today) + list(bookings_without_payment_today) + list(sessions_only_today)
    
    rows = []
    for rec in all_records:
        c_name, p_name, amount, method, p_date, b_id, total_price = rec
        cur.execute("SELECT SUM(amount), MAX(date) FROM payments WHERE booking_id = ?", (b_id,))
        pay_info = cur.fetchone()
        total_paid_ever = pay_info[0] or 0
        last_payment_date = pay_info[1]
        remaining = total_price - total_paid_ever
        display_date = last_payment_date if last_payment_date else p_date
        method_ar = "نقدي" if method == 'cash' else "محفظة" if method == 'wallet' else "انستا باي" if method == 'instapay' else method
        rows.append([c_name, p_name, amount, method_ar, remaining, display_date])
    
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التقرير اليومي"
    ws.sheet_view.rightToLeft = True
    
    headers = ["اسم العميل", "الباكيدج", "المبلغ المدفوع", "طريقة الدفع", "المبلغ المتبقي", "التاريخ"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for r in rows:
        ws.append(r)
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"daily_report_{date}.xlsx"
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@report_bp.route("/exports")
@manager_required
def exports_list():
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    files = sorted(os.listdir(EXPORTS_DIR))
    return render_template("exports.html", files=files)

@report_bp.route("/exports_download/<path:fname>")
@manager_required
def exports_download(fname):
    fpath = os.path.join(EXPORTS_DIR, fname)
    if not os.path.isfile(fpath):
        return redirect(url_for("report.exports_list"))
    return send_file(fpath, as_attachment=True)

@report_bp.route("/export_json")
@manager_required
def export_json():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name FROM employees ORDER BY id")
    employees_list = [{"id": r[0], "name": r[1]} for r in cur.fetchall()]
    cur.execute("SELECT id,name,phone,note,created_at FROM customers ORDER BY id")
    customers_list = [{"id": r[0], "name": r[1], "phone": r[2], "note": r[3], "created_at": r[4]} for r in cur.fetchall()]
    cur.execute("SELECT id,category,name,sessions_count,price,bonus FROM packages ORDER BY id")
    packages_list = [{"id": r[0], "category": r[1], "name": r[2], "sessions_count": r[3], "price": r[4], "bonus": r[5]} for r in cur.fetchall()]
    cur.execute("SELECT id,customer_id,package_id,total_sessions,sessions_done,start_date,employee_id,pulses_total,pulses_used FROM bookings ORDER BY id")
    bookings_list = [{"id": r[0], "customer_id": r[1], "package_id": r[2], "total_sessions": r[3], "sessions_done": r[4], "start_date": r[5], "employee_id": r[6], "pulses_total": r[7], "pulses_used": r[8]} for r in cur.fetchall()]
    cur.execute("SELECT id,booking_id,session_number,date,employee_id,pulses_used FROM sessions ORDER BY id")
    sessions_list = [{"id": r[0], "booking_id": r[1], "session_number": r[2], "date": r[3], "employee_id": r[4], "pulses_used": r[5]} for r in cur.fetchall()]
    conn.close()
    payload = {
        "employees": employees_list,
        "customers": customers_list,
        "packages": packages_list,
        "bookings": bookings_list,
        "sessions": sessions_list,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"pos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    fpath = os.path.join(EXPORTS_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/json")

@report_bp.route("/report_json")
@manager_required
def report_json():
    date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name FROM employees ORDER BY name")
    emps = cur.fetchall()
    rows = []
    for emp in emps:
        eid = emp[0]
        cur.execute("""SELECT COUNT(b.id), COALESCE(SUM(p.price),0) FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.employee_id=? AND b.start_date=?""", (eid, date))
        bc, bv = cur.fetchone()
        cur.execute("SELECT COUNT(id) FROM sessions WHERE employee_id=? AND date=?", (eid, date))
        sc = cur.fetchone()[0]
        rows.append({"employee": emp[1], "bookings_count": bc, "bookings_value": bv, "sessions_count": sc, "date": date})
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"daily_report_{date}.json"
    fpath = os.path.join(EXPORTS_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/json")

@report_bp.route('/export_customer/<int:customer_id>')
@manager_required
def export_customer(customer_id):
    import openpyxl
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name,phone,note,created_at FROM customers WHERE id=?", (customer_id,))
    cus = cur.fetchone()
    if not cus:
        conn.close()
        return redirect(url_for('customer_search'))
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = 'customer'
    ws_info.append(['id', 'name', 'phone', 'note', 'created_at'])
    ws_info.append(list(cus))

    ws_b = wb.create_sheet('bookings')
    ws_b.append(['id', 'package', 'price', 'total_sessions', 'sessions_done', 'start_date', 'employee'])
    cur.execute("SELECT b.id,p.name,p.price,b.total_sessions,b.sessions_done,b.start_date,b.employee_id FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.customer_id=? ORDER BY b.id", (customer_id,))
    for r in cur.fetchall():
        emp_name = None
        if r[6]:
            cur.execute('SELECT name FROM employees WHERE id=?', (r[6],))
            re = cur.fetchone()
            emp_name = re[0] if re else None
        ws_b.append([r[0], r[1], r[2], r[3], r[4], r[5], emp_name])

    ws_s = wb.create_sheet('sessions')
    ws_s.append(['id', 'booking_id', 'session_number', 'date', 'employee'])
    cur.execute('SELECT id,booking_id,session_number,date,employee_id FROM sessions WHERE booking_id IN (SELECT id FROM bookings WHERE customer_id=?) ORDER BY id', (customer_id,))
    for r in cur.fetchall():
        emp_name = None
        if r[4]:
            cur.execute('SELECT name FROM employees WHERE id=?', (r[4],))
            re = cur.fetchone()
            emp_name = re[0] if re else None
        ws_s.append([r[0], r[1], r[2], r[3], emp_name])

    ws_p = wb.create_sheet('payments')
    ws_p.append(['id', 'booking_id', 'amount', 'method', 'date', 'employee'])
    cur.execute('''SELECT p.id,p.booking_id,p.amount,p.method,p.date,p.employee_id FROM payments p WHERE p.booking_id IN (SELECT id FROM bookings WHERE customer_id=?) ORDER BY p.id''', (customer_id,))
    for r in cur.fetchall():
        emp_name = None
        if r[5]:
            cur.execute('SELECT name FROM employees WHERE id=?', (r[5],))
            re = cur.fetchone()
            emp_name = re[0] if re else None
        ws_p.append([r[0], r[1], r[2], r[3], r[4], emp_name])

    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f'customer_{customer_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

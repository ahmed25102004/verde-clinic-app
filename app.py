
from flask import Flask, render_template, request, redirect, url_for, send_file, session, flash
from datetime import datetime, timedelta
import threading
import time
import requests
import shutil
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
import json
import os
from dotenv import load_dotenv
load_dotenv()
from db import get_conn, init_db, seed_packages, EXPORTS_DIR, BACKUPS_DIR, DB_PATH
from auth import login_required, manager_required

from models import (
    Customer, Package, Booking, Session, Employee, Payment,
    Expense, Notification, WhatsAppSettings
)

app = Flask(__name__)
app.secret_key = os.environ.get("POS_SECRET", os.urandom(24))
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=os.environ.get('SESSION_COOKIE_SAMESITE', 'Lax'),
    SESSION_COOKIE_SECURE=(os.environ.get('USE_HTTPS', '0') == '1')
)

@app.template_filter('whatsapp_phone')
def whatsapp_phone_filter(phone):
    if not phone:
        return ""
    phone = str(phone).strip()
    phone = "".join(c for c in phone if c.isdigit())
    if phone.startswith('00'):
        phone = phone[2:]
    if phone.startswith('200'):
        phone = '20' + phone[3:]
    if phone.startswith('01') and len(phone) == 11:
        phone = '2' + phone
    elif phone.startswith('1') and len(phone) == 10:
        phone = '20' + phone
    return phone

def whatsapp_background_agent():
    while True:
        try:
            with app.app_context():
                settings = WhatsAppSettings.get_settings()
                if settings and settings['is_active'] == 1:
                    instance_id = settings['instance_id']
                    api_token = settings['api_token']
                    sender_phone = settings['sender_phone'] or "غير محدد"
                    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
                    
                    conn = get_conn()
                    cur = conn.cursor()
                    cur.execute('''
                        SELECT b.id, c.name, c.phone, p.name 
                        FROM bookings b
                        JOIN customers c ON b.customer_id = c.id
                        JOIN packages p ON b.package_id = p.id
                        WHERE b.next_session_date = ? AND b.reminder_sent = 0
                    ''', (tomorrow,))
                    reminders = cur.fetchall()
                    
                    for rid, cname, cphone, pname in reminders:
                        formatted_phone = whatsapp_phone_filter(cphone)
                        msg = f"مرحباً {cname}، نذكرك بموعد جلستك القادمة ({pname}) غداً في مركزنا. ننتظرك بكل حب."
                        
                        success = False
                        error_msg = ""
                        try:
                            if "instance" in instance_id:
                                url = f"https://api.ultramsg.com/{instance_id}/messages/chat"
                                payload = {"token": api_token, "to": formatted_phone, "body": msg}
                                response = requests.post(url, data=payload, timeout=10)
                                if response.status_code == 200:
                                    res_json = response.json()
                                    if res_json.get('sent') == 'true' or res_json.get('id'):
                                        success = True
                                    else:
                                        error_msg = f"UltraMsg Error: {response.text}"
                                else:
                                    error_msg = f"UltraMsg HTTP Error: {response.status_code}"
                            else:
                                status_url = f"https://api.green-api.com/waInstance{instance_id}/getStateInstance/{api_token}"
                                try:
                                    status_resp = requests.get(status_url, timeout=5)
                                    if status_resp.status_code == 200:
                                        state = status_resp.json().get('stateInstance')
                                        if state != 'authorized':
                                            error_msg = f"Green-API: الرقم غير متصل (الحالة: {state}). يرجى إعادة ربط QR Code."
                                            raise Exception(error_msg)
                                except Exception as e:
                                    if "Green-API:" in str(e):
                                        raise e
                                
                                url = f"https://api.green-api.com/waInstance{instance_id}/sendMessage/{api_token}"
                                payload = {"chatId": f"{formatted_phone}@c.us", "message": msg}
                                response = requests.post(url, json=payload, timeout=10)
                                if response.status_code == 200:
                                    res_json = response.json()
                                    if res_json.get('idMessage'):
                                        success = True
                                    else:
                                        error_msg = f"Green-API Error: {response.text}"
                                else:
                                    error_msg = f"Green-API HTTP Error: {response.status_code}"
                            
                            if success:
                                cur.execute("UPDATE bookings SET reminder_sent = 1 WHERE id = ?", (rid,))
                                cur.execute("INSERT INTO notifications (message, type) VALUES (?, ?)", 
                                           (f"تم إرسال تذكير لـ {cname} (الرقم: {formatted_phone})", "success"))
                            else:
                                cur.execute("INSERT INTO notifications (message, type) VALUES (?, ?)", 
                                           (f"فشل إرسال تذكير لـ {cname} ({formatted_phone}): {error_msg[:100]}", "danger"))
                            conn.commit()
                            time.sleep(5) 
                        except Exception as e:
                            print(f"Error sending WhatsApp to {cphone}: {e}")
                            cur.execute("INSERT INTO notifications (message, type) VALUES (?, ?)", 
                                       (f"خطأ تقني في إرسال تذكير لـ {cname}: {str(e)[:100]}", "danger"))
                            conn.commit()
                    conn.close()
        except Exception as e:
            print(f"Background Agent Error: {e}")
        time.sleep(3600)

@app.context_processor
def inject_signup_flag():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM employees')
    cnt = cur.fetchone()[0]
    conn.close()
    return {'allow_public_signup': (cnt == 0)}

# ------------------------------
# Import and register controllers
# ------------------------------
from controllers.customer_controller import (
    customer_search, customer_detail, register_customer,
    add_booking_to_customer, delete_customer, update_customer_medical
)

app.add_url_rule('/customer', 'customer_search', customer_search, methods=['GET'])
app.add_url_rule('/customer/<int:customer_id>', 'customer', customer_detail, methods=['GET'])
app.add_url_rule('/register', 'register', register_customer, methods=['GET', 'POST'])
app.add_url_rule('/customer/<int:customer_id>/add_booking', 'add_booking_existing', add_booking_to_customer, methods=['POST'])
app.add_url_rule('/customers/<int:customer_id>/delete', 'delete_customer', delete_customer, methods=['POST'])
app.add_url_rule('/customer/<int:customer_id>/update_medical', 'update_customer_medical', update_customer_medical, methods=['POST'])

# ------------------------------
# Other routes (keep as is for now, will refactor later)
# ------------------------------
@app.route("/employees", methods=["GET","POST"])
@login_required
def employees():
    conn = get_conn()
    cur = conn.cursor()
    error = None
    if request.method == "POST":
        if session.get("employee_role") != "manager":
            error = "غير مسموح: المدير فقط يستطيع إضافة موظف"
        else:
            name = request.form.get("name")
            id_str = request.form.get("id")
            pwd = request.form.get("password")
            role = request.form.get("role") or "employee"
            if not name or not id_str:
                error = "يرجى إدخال الاسم وID"
            else:
                try:
                    emp_id = int(id_str)
                    cur.execute("SELECT 1 FROM employees WHERE id=?", (emp_id,))
                    if cur.fetchone():
                        error = "المعرف مستخدم مسبقًا"
                    else:
                        ph = generate_password_hash(pwd) if pwd else None
                        cur.execute("INSERT INTO employees(id,name,password_hash,role) VALUES(?,?,?,?)", (emp_id, name, ph, role))
                        conn.commit()
                except ValueError:
                    error = "يرجى إدخال ID رقم صحيح"
    cur.execute("""SELECT e.id, e.name, e.role FROM employees e ORDER BY id DESC""")
    employees_list = cur.fetchall()
    conn.close()
    return render_template("employees.html", employees=employees_list, error=error)

@app.route("/employees/delete", methods=["POST"])
@login_required
def employees_delete():
    emp_id = request.form.get("id")
    if not emp_id:
        return redirect(url_for("employees"))
    try:
        eid = int(emp_id)
    except ValueError:
        return redirect(url_for("employees"))

    current_id = session.get("employee_id")
    role = session.get("employee_role")

    conn = get_conn()
    cur = conn.cursor()

    if eid != current_id and role != "manager":
        try:
            cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (current_id, request.remote_addr or 'unknown', 'delete_account', 0, datetime.now().isoformat()))
            conn.commit()
        except Exception:
            pass
        conn.close()
        return ("Unauthorized Access", 401)

    if eid == current_id:
        confirm_id = request.form.get("confirm_id")
        confirm_pwd = request.form.get("confirm_password")
        if not confirm_id or not confirm_pwd:
            try:
                cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (current_id, request.remote_addr or 'unknown', 'delete_account', 0, datetime.now().isoformat()))
                conn.commit()
            except Exception:
                pass
            cur.execute("SELECT id,name,role FROM employees ORDER BY id DESC")
            employees_list = cur.fetchall()
            conn.close()
            error = "بيانات غير صحيحة، لا يمكن حذف الحساب"
            return render_template("employees.html", employees=employees_list, error=error)
        try:
            confirm_eid = int(confirm_id)
        except ValueError:
            cur.execute("SELECT id,name,role FROM employees ORDER BY id DESC")
            employees_list = cur.fetchall()
            conn.close()
            error = "بيانات غير صحيحة، لا يمكن حذف الحساب"
            return render_template("employees.html", employees=employees_list, error=error)
        if confirm_eid != current_id:
            try:
                cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (current_id, request.remote_addr or 'unknown', 'delete_account', 0, datetime.now().isoformat()))
                conn.commit()
            except Exception:
                pass
            cur.execute("SELECT id,name,role FROM employees ORDER BY id DESC")
            employees_list = cur.fetchall()
            conn.close()
            error = "بيانات غير صحيحة، لا يمكن حذف الحساب"
            return render_template("employees.html", employees=employees_list, error=error)
        cur.execute("SELECT password_hash FROM employees WHERE id=?", (current_id,))
        row = cur.fetchone()
        ph = row[0] if row else None
        if not ph or not check_password_hash(ph, confirm_pwd or ""):
            try:
                cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (current_id, request.remote_addr or 'unknown', 'delete_account', 0, datetime.now().isoformat()))
                conn.commit()
            except Exception:
                pass
            cur.execute("SELECT id,name,role FROM employees ORDER BY id DESC")
            employees_list = cur.fetchall()
            conn.close()
            error = "بيانات غير صحيحة، لا يمكن حذف الحساب"
            return render_template("employees.html", employees=employees_list, error=error)

    cur.execute("UPDATE bookings SET employee_id=NULL WHERE employee_id=?", (eid,))
    cur.execute("UPDATE sessions SET employee_id=NULL WHERE employee_id=?", (eid,))
    cur.execute("DELETE FROM employees WHERE id=?", (eid,))
    try:
        cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (eid if eid == current_id else current_id, request.remote_addr or 'unknown', 'delete_account', 1, datetime.now().isoformat()))
    except Exception:
        pass
    conn.commit()
    conn.close()

    if eid == current_id:
        session.clear()
        return redirect(url_for("signup"))
    return redirect(url_for("employees"))

@app.route("/")
def index():
    if not session.get("employee_id"):
        return redirect(url_for("signup"))
    cosmetic_sessions = Package.find_by_category("cosmetic_sessions")
    cosmetic_packages = Package.find_by_category("cosmetic_packages")
    laser_sessions = Package.find_by_category("laser_sessions")
    laser_packages = Package.find_by_category("laser_packages")
    pulse_packages = Package.find_by_category("pulse_packages")
    return render_template("index.html", 
                         cosmetic_sessions=[(p['id'], p['name'], p['price']) for p in cosmetic_sessions],
                         cosmetic_packages=[(p['id'], p['name'], p['price']) for p in cosmetic_packages],
                         laser_sessions=[(p['id'], p['name'], p['price']) for p in laser_sessions],
                         laser_packages=[(p['id'], p['name'], p['price']) for p in laser_packages],
                         pulse_packages=[(p['id'], p['name'], p['price']) for p in pulse_packages])

@app.route('/check_phone')
@login_required
def check_phone():
    phone = (request.args.get('phone') or '').strip()
    if not phone:
        return app.response_class(json.dumps({'exists': False}), 200, {'Content-Type': 'application/json'})
    if not phone.isdigit() or len(phone) != 11:
        return app.response_class(json.dumps({'error': 'رقم الموبايل يجب أن يكون 11 رقمًا'}), 400, {'Content-Type': 'application/json'})
    return app.response_class(json.dumps({'exists': False}), 200, {'Content-Type': 'application/json'})

@app.route("/bookings/<int:booking_id>/update_next_session", methods=["POST"])
def update_next_session(booking_id):
    if not session.get("employee_id"):
        return redirect(url_for("login"))
    next_date = request.form.get("next_session_date")
    Booking.update_next_session_date(booking_id, next_date)
    flash("تم تحديث موعد الجلسة القادمة بنجاح", "success")
    return redirect(request.referrer or url_for("index"))

@app.route("/bookings/<int:booking_id>/add_session", methods=["POST"])
def add_session(booking_id):
    if not session.get("employee_id"):
        return redirect(url_for("login"))
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT total_sessions,sessions_done,pulses_total,pulses_used FROM bookings WHERE id=?", (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return redirect(url_for("index"))
    total, done, pulses_total, pulses_used = row
    today = datetime.now().strftime("%Y-%m-%d")
    
    if session.get('employee_role') == 'manager':
        employee_id = request.form.get('employee_id') or session.get('employee_id')
    else:
        employee_id = session.get('employee_id')
    try:
        employee_id = int(employee_id)
    except Exception:
        employee_id = session.get('employee_id')
    
    if pulses_total and pulses_total > 0:
        try:
            use = int(request.form.get('pulses_used') or 0)
        except Exception:
            use = 0
        if use <= 0:
            conn.close()
            return ("عدد النبضات غير كافٍ", 400)
        if pulses_used + use > pulses_total:
            conn.close()
            return ("عدد النبضات غير كافٍ", 400)
        note_val = request.form.get('note')
        Session.create(booking_id, done + 1, today, employee_id, use, note_val)
        cur.execute("UPDATE bookings SET pulses_used = pulses_used + ? WHERE id = ?", (use, booking_id))
    else:
        if done < total:
            session_number = done + 1
            note_val = request.form.get('note')
            Session.create(booking_id, session_number, today, employee_id, 0, note_val)
            cur.execute("UPDATE bookings SET sessions_done = sessions_done + 1 WHERE id = ?", (booking_id,))
        else:
            conn.close()
            return redirect(request.referrer or url_for("index"))
    
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for("index"))

@app.route("/bookings/<int:booking_id>/pay", methods=["POST"])
def add_payment(booking_id):
    if not session.get("employee_id"):
        return redirect(url_for("login"))
    method = (request.form.get('method') or '').strip().lower()
    amount_str = request.form.get('amount')
    allowed = {"cash","wallet","instapay"}
    try:
        amount = int(amount_str or '0')
    except Exception:
        amount = 0
    
    if method not in allowed or amount <= 0:
        flash("بيانات الدفع غير صحيحة", "danger")
        return redirect(request.referrer or url_for("index"))
    
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''SELECT b.price_override, p.price, (SELECT COALESCE(SUM(pay.amount), 0) FROM payments pay WHERE pay.booking_id = b.id) as total_paid FROM bookings b JOIN packages p ON p.id = b.package_id WHERE b.id = ?''', (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        flash("الحجز غير موجود", "danger")
        return redirect(url_for("index"))
    price_override, package_price, total_paid = row
    actual_price = price_override if price_override is not None else package_price
    remaining = actual_price - total_paid
    
    if amount > remaining:
        conn.close()
        flash(f"لا يمكن دفع مبلغ أكبر من المتبقي ({remaining} جنيه)", "danger")
        return redirect(request.referrer or url_for("index"))
    
    today = datetime.now().strftime("%Y-%m-%d")
    Payment.create(booking_id, amount, method, today, session.get('employee_id'))
    flash("تم تسجيل عملية الدفع بنجاح", "success")
    return redirect(request.referrer or url_for("index"))

@app.route("/sessions/<int:sid>/delete", methods=["POST"])
@manager_required
def delete_session(sid):
    booking_id = Session.delete_and_update_booking(sid)
    return redirect(request.referrer or url_for("index"))

@app.route("/sessions/<int:sid>/update_note", methods=["POST"])
@manager_required
def update_session_note(sid):
    note = request.form.get('note') or ''
    Session.update_note(sid, note)
    return redirect(request.referrer or url_for("index"))

@app.route("/bookings/<int:bid>/delete", methods=["POST"])
@manager_required
def delete_booking(bid):
    customer_id = Booking.delete_with_relations(bid)
    if customer_id:
        return redirect(url_for("customer", customer_id=customer_id))
    else:
        return redirect(request.referrer or url_for("index"))

@app.route('/report')
@manager_required
def report_daily():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date_single = request.args.get('date')
    
    if date_single:
        start_date = date_single
        end_date = date_single
    
    if not start_date:
        start_date = datetime.now().strftime('%Y-%m-%d')
    if not end_date:
        end_date = start_date

    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute('SELECT id,name FROM employees ORDER BY name')
    emps = cur.fetchall()
    
    expenses_map = Expense.get_expenses_map_by_date_range(start_date, end_date)
    
    rows = []
    total_cash = 0
    total_wallet = 0
    total_instapay = 0
    total_expenses_total = 0
    
    for emp in emps:
        eid = emp[0]
        cur.execute('SELECT method, COALESCE(SUM(amount),0) FROM payments WHERE employee_id=? AND date BETWEEN ? AND ? GROUP BY method', (eid, start_date, end_date))
        pms = cur.fetchall()
        pm = {'cash': 0, 'wallet': 0, 'instapay': 0}
        for m in pms:
            if m[0] in pm:
                pm[m[0]] = m[1]
        
        emp_expenses = expenses_map.get(eid, 0)
        total_for_manager = (pm['cash'] + pm['wallet'] + pm['instapay']) - emp_expenses
        
        total_cash += pm['cash']
        total_wallet += pm['wallet']
        total_instapay += pm['instapay']
        total_expenses_total += emp_expenses
        
        cur.execute('SELECT COUNT(id) FROM sessions WHERE employee_id=? AND date BETWEEN ? AND ?', (eid, start_date, end_date))
        sessions_count = cur.fetchone()[0]
        
        cur.execute('''SELECT COUNT(b.id), COALESCE(SUM(p.price),0) FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.employee_id=? AND b.start_date BETWEEN ? AND ?''', (eid, start_date, end_date))
        bc, bv = cur.fetchone()
        
        rows.append({
            'employee': emp[1], 'employee_id': eid, 'cash': pm['cash'], 'wallet': pm['wallet'],
            'instapay': pm['instapay'], 'total_payments': pm['cash'] + pm['wallet'] + pm['instapay'],
            'expenses': emp_expenses, 'total_for_manager': total_for_manager, 'sessions_count': sessions_count,
            'bookings_count': bc or 0, 'bookings_value': bv or 0
        })
    
    grand_total = total_cash + total_wallet + total_instapay
    net_total = grand_total - total_expenses_total
    
    cur.execute('SELECT COALESCE(SUM(amount),0) FROM payments WHERE date BETWEEN ? AND ?', (start_date, end_date))
    payments_total = cur.fetchone()[0]
    
    cur.execute('SELECT method, COALESCE(SUM(amount),0) FROM payments WHERE date BETWEEN ? AND ? GROUP BY method', (start_date, end_date))
    payments_by_method = cur.fetchall()
    
    cur.execute('SELECT COUNT(*) FROM payments WHERE date BETWEEN ? AND ?', (start_date, end_date))
    payments_count = cur.fetchone()[0]
    
    cur.execute('SELECT COUNT(DISTINCT b.customer_id) FROM payments p JOIN bookings b ON b.id=p.booking_id WHERE p.date BETWEEN ? AND ?', (start_date, end_date))
    unique_payers = cur.fetchone()[0]
    
    cur.execute('SELECT COALESCE(SUM(p.price),0), COUNT(b.id) FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.start_date BETWEEN ? AND ?', (start_date, end_date))
    bv_row = cur.fetchone()
    bookings_value_today = bv_row[0] if bv_row else 0
    bookings_count_today = bv_row[1] if bv_row else 0
    
    cur.execute('SELECT COUNT(*) FROM sessions WHERE date BETWEEN ? AND ?', (start_date, end_date))
    sessions_count_today = cur.fetchone()[0]
    
    cur.execute('''SELECT COALESCE(SUM(amount), 0) FROM payments WHERE booking_id IN (SELECT id FROM bookings WHERE start_date BETWEEN ? AND ?)''', (start_date, end_date))
    bookings_paid_sum = cur.fetchone()[0]
    bookings_remaining = bookings_value_today - bookings_paid_sum

    conn.close()
    
    summary = {
        'start_date': start_date, 'end_date': end_date, 'total_cash': total_cash, 'total_wallet': total_wallet,
        'total_instapay': total_instapay, 'total_expenses': total_expenses_total, 'grand_total': grand_total,
        'net_total': net_total, 'payments_total': payments_total or 0, 'payments_by_method': payments_by_method or [],
        'payments_count': payments_count or 0, 'unique_payers': unique_payers or 0, 'bookings_value_today': bookings_value_today or 0,
        'bookings_count_today': bookings_count_today or 0, 'sessions_count_today': sessions_count_today or 0,
        'bookings_paid_sum': bookings_paid_sum, 'bookings_remaining': bookings_remaining
    }
    return render_template('report.html', rows=rows, summary=summary, start_date=start_date, end_date=end_date)

@app.route('/dashboard')
@manager_required
def dashboard():
    conn = get_conn()
    cur = conn.cursor()
    
    last_7_days = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        cur.execute('SELECT COALESCE(SUM(amount), 0) FROM payments WHERE date=?', (d,))
        last_7_days.append({'date': d, 'amount': cur.fetchone()[0]})
    
    cur.execute('''SELECT e.name, (SELECT COUNT(*) FROM sessions WHERE employee_id = e.id) as sessions, (SELECT COALESCE(SUM(amount), 0) FROM payments WHERE employee_id = e.id) as revenue FROM employees e ORDER BY revenue DESC''')
    employee_stats = cur.fetchall()
    
    cur.execute('''SELECT p.name, COUNT(b.id) as bookings_count FROM packages p LEFT JOIN bookings b ON p.id = b.package_id GROUP BY p.id ORDER BY bookings_count DESC LIMIT 5''')
    top_packages = cur.fetchall()
    
    cur.execute('''SELECT c.name, COALESCE(SUM(p.amount), 0) as total_paid FROM customers c JOIN bookings b ON c.id = b.customer_id JOIN payments p ON b.id = p.booking_id GROUP BY c.id ORDER BY total_paid DESC LIMIT 5''')
    top_customers = cur.fetchall()

    cur.execute("SELECT message, created_at FROM notifications WHERE message LIKE 'تم إرسال تذكير تلقائي%' ORDER BY created_at DESC LIMIT 5")
    whatsapp_logs = cur.fetchall()

    insights = []
    tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')
    today_str = datetime.now().strftime('%Y-%m-%d')
    cur.execute('''SELECT c.name, c.phone, b.next_session_date, p.name, b.id FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages p ON b.package_id = p.id WHERE b.next_session_date = ? OR b.next_session_date = ?''', (today_str, tomorrow))
    upcoming_appointments = cur.fetchall()
    
    for appt in upcoming_appointments:
        day_text = "اليوم" if appt[2] == today_str else "غداً"
        phone = appt[1] or ""
        if phone.startswith("0"):
            phone = "2" + phone
        elif len(phone) == 10 and phone.startswith("1"):
            phone = "20" + phone
            
        insights.append({
            'type': 'primary', 'text': f"موعد جلسة {appt[0]} ({appt[3]}) {day_text}",
            'phone': phone, 'action': 'whatsapp',
            'msg': f"مرحباً {appt[0]}، نذكرك بموعد جلستك القادمة ({appt[3]}) {day_text} في مركزنا. ننتظرك بكل حب."
        })

    cur.execute('''SELECT c.name, (COALESCE(b.price_override, pkg.price) - (SELECT COALESCE(SUM(amount), 0) FROM payments WHERE booking_id = b.id)) as remaining FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE remaining > 500 ORDER BY remaining DESC LIMIT 3''')
    debtors = cur.fetchall()
    for d in debtors:
        insights.append({'type': 'warning', 'text': f"العميلة {d[0]} عليها مبلغ متبقي كبير ({d[1]} ج.م). يُنصح بالمتابعة لتحصيل المبلغ."})

    cur.execute('''SELECT c.name, pkg.name FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE b.sessions_done >= b.total_sessions AND NOT EXISTS (SELECT 1 FROM bookings b2 WHERE b2.customer_id = b.customer_id AND b2.id > b.id) LIMIT 3''')
    finished = cur.fetchall()
    for f in finished:
        insights.append({'type': 'info', 'text': f"العميلة {f[0]} أكملت جميع جلسات {f[1]}. يمكنك التواصل معها لعرض باكدج جديد."})

    today = datetime.now().strftime('%Y-%m-%d')
    cur.execute('SELECT COUNT(*) FROM sessions WHERE date = ?', (today,))
    today_sessions = cur.fetchone()[0]
    if today_sessions == 0:
        insights.append({'type': 'danger', 'text': "لم يتم تسجيل أي جلسات اليوم حتى الآن. تأكد من سير العمل بشكل طبيعي."})

    conn.close()
    return render_template('dashboard.html', 
                         last_7_days=last_7_days, 
                         employee_stats=employee_stats,
                         top_packages=top_packages,
                         top_customers=top_customers,
                         insights=insights,
                         whatsapp_logs=whatsapp_logs)

@app.route("/quick_session", methods=["GET","POST"])
def quick_session():
    if not session.get("employee_id"):
        return redirect(url_for("login", next=request.path))
    error = None
    if request.method == "POST":
        booking_id = request.form.get("booking_id")
        if session.get('employee_role') == 'manager':
            employee_id = request.form.get('employee_id') or session.get('employee_id')
        else:
            employee_id = session.get('employee_id')
        if not booking_id or not employee_id:
            error = "يرجى إدخال رقم العملية ورقم الموظف"
        else:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT customer_id,total_sessions,sessions_done FROM bookings WHERE id=?", (booking_id,))
            b = cur.fetchone()
            if not b:
                error = "لا يوجد عملية بهذا الرقم"
            else:
                customer_id, total, done = b
                if done >= total:
                    error = "هذا الحجز مكتمل"
                else:
                    sn = done + 1
                    today = datetime.now().strftime("%Y-%m-%d")
                    note_val = request.form.get('note')
                    Session.create(booking_id, sn, today, employee_id, 0, note_val)
                    cur.execute("UPDATE bookings SET sessions_done = sessions_done + 1 WHERE id = ?", (booking_id,))
                    conn.commit()
                    conn.close()
                    return redirect(url_for("customer", id=customer_id))
            if conn:
                conn.close()
    return render_template("quick_session.html", error=error)

@app.route("/export")
@manager_required
def export_all():
    import openpyxl
    wb = openpyxl.Workbook()
    ws_emp = wb.active
    ws_emp.title = "employees"
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name FROM employees ORDER BY id")
    ws_emp.append(["id","name"])
    for r in cur.fetchall():
        ws_emp.append(list(r))
    ws_cus = wb.create_sheet("customers")
    ws_cus.append(["id","name","phone","note","created_at"])
    cur.execute("SELECT id,name,phone,note,created_at FROM customers ORDER BY id")
    for r in cur.fetchall():
        ws_cus.append(list(r))
    ws_pkg = wb.create_sheet("packages")
    ws_pkg.append(["id","category","name","sessions_count","price","bonus"])
    cur.execute("SELECT id,category,name,sessions_count,price,bonus FROM packages ORDER BY id")
    for r in cur.fetchall():
        ws_pkg.append(list(r))
    ws_b = wb.create_sheet("bookings")
    ws_b.append(["id","customer_id","package_id","total_sessions","sessions_done","start_date","employee_id","pulses_total","pulses_used"])
    cur.execute("SELECT id,customer_id,package_id,total_sessions,sessions_done,start_date,employee_id,pulses_total,pulses_used FROM bookings ORDER BY id")
    for r in cur.fetchall():
        ws_b.append(list(r))
    ws_s = wb.create_sheet("sessions")
    ws_s.append(["id","booking_id","session_number","date","employee_id","pulses_used"])
    cur.execute("SELECT id,booking_id,session_number,date,employee_id,pulses_used FROM sessions ORDER BY id")
    for r in cur.fetchall():
        ws_s.append(list(r))
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"pos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/report_export")
@manager_required
def report_export():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    start_date = request.args.get("start_date") or datetime.now().strftime("%Y-%m-%d")
    end_date = request.args.get("end_date") or start_date
    
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("""SELECT c.name, pkg.category, pkg.name, p.amount, p.date, p.method FROM payments p JOIN bookings b ON p.booking_id = b.id JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE p.date BETWEEN ? AND ? ORDER BY p.date ASC, p.id ASC""", (start_date, end_date))
    payments = cur.fetchall()
    
    cur.execute("""SELECT method, COALESCE(SUM(amount), 0) FROM payments WHERE date BETWEEN ? AND ? GROUP BY method""", (start_date, end_date))
    method_totals = {row[0]: row[1] for row in cur.fetchall()}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.rightToLeft = True
    
    headers = ["اسم العميل", "الفئة (Area)", "الباكدج", "المبلغ المدفوع", "التاريخ", "طريقة الدفع"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for p in payments:
        method_ar = "نقدي" if p[5] == 'cash' else "محفظة" if p[5] == 'wallet' else "انستا باي" if p[5] == 'instapay' else p[5]
        ws.append([p[0], p[1], p[2], p[3], p[4], method_ar])
    
    ws.append([])
    
    ws.append(["إجمالي النقدي", method_totals.get('cash', 0), "جنيه"])
    ws.append(["إجمالي المحفظة", method_totals.get('wallet', 0), "جنيه"])
    ws.append(["إجمالي انستا باي", method_totals.get('instapay', 0), "جنيه"])
    ws.append(["الإجمالي الكلي", sum(method_totals.values()), "جنيه"])
    
    last_row = ws.max_row
    for i in range(last_row - 3, last_row + 1):
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.cell(row=i, column=2).font = Font(bold=True)
    
    conn.close()
    
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"report_{start_date}_to_{end_date}.xlsx"
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/export_daily_customers")
@manager_required
def export_daily_customers():
    import openpyxl
    from openpyxl.styles import Font, Alignment
    
    date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    conn = get_conn()
    cur = conn.cursor()
    
    cur.execute("""SELECT c.name as customer_name, pkg.name as package_name, p.amount as paid_amount, p.method as payment_method, p.date as transaction_date, b.id as booking_id, COALESCE(b.price_override, pkg.price) as total_price FROM payments p JOIN bookings b ON p.booking_id = b.id JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE p.date = ?""", (date,))
    payments_today = cur.fetchall()
    
    cur.execute("""SELECT c.name as customer_name, pkg.name as package_name, 0 as paid_amount, 'لم يدفع' as payment_method, b.start_date as transaction_date, b.id as booking_id, COALESCE(b.price_override, pkg.price) as total_price FROM bookings b JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE b.start_date = ? AND b.id NOT IN (SELECT booking_id FROM payments WHERE date = ?)""", (date, date))
    bookings_without_payment_today = cur.fetchall()

    cur.execute("""SELECT DISTINCT c.name as customer_name, pkg.name as package_name, 0 as paid_amount, 'لم يدفع' as payment_method, b.start_date as transaction_date, b.id as booking_id, COALESCE(b.price_override, pkg.price) as total_price FROM sessions s JOIN bookings b ON s.booking_id = b.id JOIN customers c ON b.customer_id = c.id JOIN packages pkg ON b.package_id = pkg.id WHERE s.date = ? AND b.id NOT IN (SELECT booking_id FROM payments WHERE date = ?) AND b.start_date != ?""", (date, date, date))
    sessions_only_today = cur.fetchall()
    
    all_records = list(payments_today) + list(bookings_without_payment_today) + list(sessions_only_today)
    
    rows = []
    for rec in all_records:
        c_name, p_name, amount, method, p_date, b_id, total_price = rec
        
        cur.execute("SELECT SUM(amount), MAX(date) FROM payments WHERE booking_id = ?", (b_id,))
        pay_info = cur.fetchone()
        total_paid_ever = pay_info[0] or 0
        last_payment_date = pay_info[1]
        
        remaining = total_price - total_paid_ever
        
        display_date = last_payment_date if last_payment_date else p_date
        
        method_ar = "نقدي" if method == 'cash' else "محفظة" if method == 'wallet' else "انستا باي" if method == 'instapay' else method
        
        rows.append([c_name, p_name, amount, method_ar, remaining, display_date])
    
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التقرير اليومي"
    ws.sheet_view.rightToLeft = True
    
    headers = ["اسم العميل", "الباكيدج", "المبلغ المدفوع", "طريقة الدفع", "المبلغ المتبقي", "التاريخ"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for r in rows:
        ws.append(r)
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"daily_report_{date}.xlsx"
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    
    return send_file(fpath, as_attachment=True, download_name=fpath, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/exports")
@manager_required
def exports_list():
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    files = sorted(os.listdir(EXPORTS_DIR))
    return render_template("exports.html", files=files)

@app.route("/exports_download/<path:fname>")
@manager_required
def exports_download(fname):
    fpath = os.path.join(EXPORTS_DIR, fname)
    if not os.path.isfile(fpath):
        return redirect(url_for("exports_list"))
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route("/backups")
@manager_required
def backups_list():
    os.makedirs(BACKUPS_DIR, exist_ok=True)
    files = sorted([f for f in os.listdir(BACKUPS_DIR) if f.endswith('.db')], reverse=True)
    return render_template("backups.html", files=files)

@app.route("/backups/create")
@manager_required
def backup_create():
    flash("النسخ الاحتياطي لـ SQL Server يتم إدارته من خلال SQL Server Management Studio أو أوامر BACKUP DATABASE.", "info")
    return redirect(url_for("backups_list"))

@app.route("/backups/download/<path:fname>")
@manager_required
def backup_download(fname):
    fpath = os.path.join(BACKUPS_DIR, fname)
    if not os.path.isfile(fpath):
        flash("الملف غير موجود", "danger")
        return redirect(url_for("backups_list"))
    return send_file(fpath, as_attachment=True, download_name=fname)

@app.route("/backups/delete/<path:fname>")
@manager_required
def backup_delete(fname):
    fpath = os.path.join(BACKUPS_DIR, fname)
    if os.path.isfile(fpath):
        try:
            os.remove(fpath)
            flash("تم حذف النسخة الاحتياطية بنجاح", "success")
        except Exception as e:
            flash(f"خطأ في الحذف: {e}", "danger")
    return redirect(url_for("backups_list"))

@app.route("/backups/restore/<path:fname>")
@manager_required
def backup_restore(fname):
    flash("استعادة قاعدة بيانات SQL Server يتم إدارته من خلال SQL Server Management Studio أو أوامر RESTORE DATABASE.", "info")
    return redirect(url_for("backups_list"))

@app.route("/customers")
@manager_required
def customers_list():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""SELECT c.id, c.name, c.phone, c.note, c.created_at, COUNT(b.id) AS bookings_count FROM customers c LEFT JOIN bookings b ON b.customer_id = c.id GROUP BY c.id ORDER BY c.id DESC""")
    rows = cur.fetchall()
    conn.close()
    return render_template("customers.html", rows=rows)

@app.context_processor
def inject_notifications():
    if session.get('employee_role') == 'manager':
        try:
            unread_notifs = Notification.find_unread()
            return {'notifications': [(n['message'], n['type'], n['created_at']) for n in unread_notifs], 'unread_count': len(unread_notifs)}
        except Exception:
            return {'notifications': [], 'unread_count': 0}
    return {'notifications': [], 'unread_count': 0}

@app.route("/notifications/read_all")
@manager_required
def read_all_notifications():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE notifications SET is_read = 1")
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('dashboard'))

@app.route("/login", methods=["GET","POST"])
def login():
    error = None
    if request.method == "POST":
        id_str = request.form.get("id")
        pwd = request.form.get("password")
        next_url = request.args.get("next") or url_for("index")
        try:
            emp_id = int(id_str)
        except (TypeError, ValueError):
            error = "ID غير صالح"
        else:
            ip = request.remote_addr or 'unknown'
            conn = get_conn()
            cur = conn.cursor()
            cur.execute('SELECT attempts,last_attempt FROM auth_failures WHERE ip=?', (ip,))
            af = cur.fetchone()
            locked = False
            if af:
                try:
                    last = datetime.fromisoformat(af[1])
                except Exception:
                    last = datetime.now()
                if af[0] >= 5 and (datetime.now() - last).total_seconds() < 15*60:
                    locked = True
            if locked:
                error = 'محظور مؤقتًا بعد محاولات فاشلة، حاول بعد قليل'
            else:
                cur.execute("SELECT id,name,password_hash,role FROM employees WHERE id=?", (emp_id,))
                row = cur.fetchone()
                if not row or not row[2] or not check_password_hash(row[2], pwd or ""):
                    error = "بيانات تسجيل الدخول غير صحيحة"
                    now_s = datetime.now().isoformat()
                    try:
                        cur.execute('INSERT INTO auth_failures(ip,attempts,last_attempt) VALUES(?,?,?)', (ip, 1, now_s))
                    except Exception:
                        cur.execute('UPDATE auth_failures SET attempts=attempts+1,last_attempt=? WHERE ip=?', (now_s, ip))
                    cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (emp_id, ip, 'login', 0, datetime.now().isoformat()))
                    conn.commit()
                else:
                    cur.execute('DELETE FROM auth_failures WHERE ip=?', (ip,))
                    session["employee_id"] = row[0]
                    session["employee_name"] = row[1]
                    session["employee_role"] = row[3] or "employee"
                    
                    if session["employee_role"] == "manager":
                        today_str = datetime.now().strftime('%Y-%m-%d')
                        backup_file = os.path.join(BACKUPS_DIR, f"verde_clinic_backup_{today_str}.db")
                        if not os.path.exists(backup_file):
                            try:
                                os.makedirs(BACKUPS_DIR, exist_ok=True)
                                shutil.copy2(DB_PATH, backup_file)
                                for f in os.listdir(BACKUPS_DIR):
                                    f_path = os.path.join(BACKUPS_DIR, f)
                                    if os.path.isfile(f_path):
                                        f_date_str = f.replace('verde_clinic_backup_', '').replace('.db', '')
                                        try:
                                            f_date = datetime.strptime(f_date_str, '%Y-%m-%d')
                                            if (datetime.now() - f_date).days > 7:
                                                os.remove(f_path)
                                        except: pass
                            except Exception: pass

                    cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (row[0], ip, 'login', 1, datetime.now().isoformat()))
                    conn.commit()
                    conn.close()
                    return redirect(next_url)
            conn.commit()
            conn.close()
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    employee_id = session.get("employee_id")
    conn = get_conn()
    cur = conn.cursor()
    
    if request.method == "POST":
        current_pwd = request.form.get("current_password")
        new_pwd = request.form.get("new_password")
        confirm_pwd = request.form.get("confirm_password")
        
        cur.execute("SELECT password_hash FROM employees WHERE id = ?", (employee_id,))
        user = cur.fetchone()
        
        if not user or not check_password_hash(user[0], current_pwd):
            flash("كلمة المرور الحالية غير صحيحة", "danger")
        elif not new_pwd or len(new_pwd) < 8:
            flash("كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل", "danger")
        elif new_pwd != confirm_pwd:
            flash("تأكيد كلمة المرور لا يطابق كلمة المرور الجديدة", "danger")
        else:
            new_hash = generate_password_hash(new_pwd)
            cur.execute("UPDATE employees SET password_hash = ? WHERE id = ?", (new_hash, employee_id))
            conn.commit()
            flash("تم تغيير كلمة المرور بنجاح", "success")
            
    cur.execute("SELECT id, name, role FROM employees WHERE id = ?", (employee_id,))
    employee = cur.fetchone()
    
    cur.execute("SELECT COUNT(*) FROM sessions WHERE employee_id = ?", (employee_id,))
    sessions_count = cur.fetchone()[0]
    
    cur.execute("SELECT COALESCE(SUM(amount), 0) FROM payments WHERE employee_id = ?", (employee_id,))
    total_collected = cur.fetchone()[0]
    
    month_ago = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    cur.execute("SELECT COUNT(*) FROM sessions WHERE employee_id = ? AND date >= ?", (employee_id, month_ago))
    sessions_month = cur.fetchone()[0]
    
    cur.execute("SELECT COALESCE(SUM(amount), 0) FROM payments WHERE employee_id = ? AND date >= ?", (employee_id, month_ago))
    collected_month = cur.fetchone()[0]
    
    conn.close()
    
    return render_template("profile.html", 
                         employee=employee, 
                         sessions_count=sessions_count, 
                         total_collected=total_collected,
                         sessions_month=sessions_month,
                         collected_month=collected_month)

@app.route('/signup', methods=['GET','POST'])
def signup():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM employees')
    cnt = cur.fetchone()[0]
    allow_manager = (cnt > 0)
    error = None
    if request.method == 'POST':
        if not (cnt == 0 or session.get('employee_role') == 'manager'):
            error = 'غير مسموح: تحتاج لتكون مديرًا لتنشئ حسابًا جديدًا'
        else:
            id_str = request.form.get('id')
            name = request.form.get('name')
            pwd = request.form.get('password')
            pwd2 = request.form.get('password_confirm')
            if not pwd or len(pwd) < 8:
                error = 'كلمة المرور قصيرة جدًا (8 أحرف على الأقل)'
            elif pwd != (pwd2 or ''):
                error = 'تأكيد كلمة المرور لا يطابق'
            else:
                role = request.form.get('role') or 'employee'
                try:
                    eid = int(id_str)
                except (TypeError, ValueError):
                    error = 'ID غير صالح'
                else:
                    cur.execute('SELECT 1 FROM employees WHERE id=?', (eid,))
                    if cur.fetchone():
                        error = 'هذا المعرف مستخدم بالفعل'
                    elif not name or not pwd:
                        error = 'يرجى إدخال الاسم وكلمة المرور'
                    else:
                        ph = generate_password_hash(pwd)
                        cur.execute('INSERT INTO employees(id,name,password_hash,role) VALUES(?,?,?,?)', (eid, name, ph, role))
                        cur.execute('INSERT INTO auth_logs(employee_id,ip,action,success,timestamp) VALUES(?,?,?,?,?)', (eid, request.remote_addr or 'unknown', 'create_account', 1, datetime.now().isoformat()))
                        conn.commit()
                        conn.close()
                        return redirect(url_for('login'))
    conn.close()
    return render_template('signup.html', error=error, allow_manager=allow_manager)

@app.route("/packages", methods=["GET","POST"])
@manager_required
def packages_admin():
    conn = get_conn()
    cur = conn.cursor()
    error = None
    if request.method == "POST":
        category = request.form.get("category")
        name = request.form.get("name")
        sessions_count = request.form.get("sessions_count")
        price = request.form.get("price")
        bonus = request.form.get("bonus")
        allowed = {"cosmetic_sessions","cosmetic_packages","laser_sessions","laser_packages","pulse_packages"}
        try:
            sc = int(sessions_count or 0)
            pr = int(price or 0)
            if category not in allowed or not name or sc <= 0 or pr <= 0:
                error = "بيانات غير صحيحة"
            else:
                cur.execute("INSERT INTO packages(category,name,sessions_count,price,bonus) VALUES(?,?,?,?,?)", (category, name, sc, pr, bonus))
                conn.commit()
        except ValueError:
            error = "القيم الرقمية غير صحيحة"
    cur.execute("SELECT id,category,name,sessions_count,price FROM packages ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return render_template("packages.html", rows=rows, error=error)

@app.route("/packages/delete", methods=["POST"])
@manager_required
def packages_delete():
    pid = request.form.get("id")
    try:
        p = int(pid)
    except (TypeError, ValueError):
        return redirect(url_for("packages_admin"))
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM packages WHERE id=?", (p,))
    conn.commit()
    conn.close()
    return redirect(url_for("packages_admin"))

@app.route("/export_json")
@manager_required
def export_json():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name FROM employees ORDER BY id")
    employees_list = [{"id": r[0], "name": r[1]} for r in cur.fetchall()]
    cur.execute("SELECT id,name,phone,note,created_at FROM customers ORDER BY id")
    customers_list = [{"id": r[0], "name": r[1], "phone": r[2], "note": r[3], "created_at": r[4]} for r in cur.fetchall()]
    cur.execute("SELECT id,category,name,sessions_count,price,bonus FROM packages ORDER BY id")
    packages_list = [{"id": r[0], "category": r[1], "name": r[2], "sessions_count": r[3], "price": r[4], "bonus": r[5]} for r in cur.fetchall()]
    cur.execute("SELECT id,customer_id,package_id,total_sessions,sessions_done,start_date,employee_id,pulses_total,pulses_used FROM bookings ORDER BY id")
    bookings_list = [{"id": r[0], "customer_id": r[1], "package_id": r[2], "total_sessions": r[3], "sessions_done": r[4], "start_date": r[5], "employee_id": r[6], "pulses_total": r[7], "pulses_used": r[8]} for r in cur.fetchall()]
    cur.execute("SELECT id,booking_id,session_number,date,employee_id,pulses_used FROM sessions ORDER BY id")
    sessions_list = [{"id": r[0], "booking_id": r[1], "session_number": r[2], "date": r[3], "employee_id": r[4], "pulses_used": r[5]} for r in cur.fetchall()]
    conn.close()
    payload = {
        "employees": employees_list,
        "customers": customers_list,
        "packages": packages_list,
        "bookings": bookings_list,
        "sessions": sessions_list,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"pos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    fpath = os.path.join(EXPORTS_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/json")

@app.route("/report_json")
@manager_required
def report_json():
    date = request.args.get("date") or datetime.now().strftime("%Y-%m-%d")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name FROM employees ORDER BY name")
    emps = cur.fetchall()
    rows = []
    for emp in emps:
        eid = emp[0]
        cur.execute("""SELECT COUNT(b.id), COALESCE(SUM(p.price),0) FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.employee_id=? AND b.start_date=?""", (eid, date))
        bc, bv = cur.fetchone()
        cur.execute("SELECT COUNT(id) FROM sessions WHERE employee_id=? AND date=?", (eid, date))
        sc = cur.fetchone()[0]
        rows.append({"employee": emp[1], "bookings_count": bc, "bookings_value": bv, "sessions_count": sc, "date": date})
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f"daily_report_{date}.json"
    fpath = os.path.join(EXPORTS_DIR, fname)
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype="application/json")

@app.route('/export_customer/<int:customer_id>')
@manager_required
def export_customer(customer_id):
    import openpyxl
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id,name,phone,note,created_at FROM customers WHERE id=?", (customer_id,))
    cus = cur.fetchone()
    if not cus:
        conn.close()
        return redirect(url_for('customer_search'))
    wb = openpyxl.Workbook()
    ws_info = wb.active
    ws_info.title = 'customer'
    ws_info.append(['id','name','phone','note','created_at'])
    ws_info.append(list(cus))

    ws_b = wb.create_sheet('bookings')
    ws_b.append(['id','package','price','total_sessions','sessions_done','start_date','employee'])
    cur.execute("SELECT b.id,p.name,p.price,b.total_sessions,b.sessions_done,b.start_date,b.employee_id FROM bookings b JOIN packages p ON p.id=b.package_id WHERE b.customer_id=? ORDER BY b.id", (customer_id,))
    for r in cur.fetchall():
        emp_name = None
        if r[6]:
            cur.execute('SELECT name FROM employees WHERE id=?', (r[6],))
            re = cur.fetchone()
            emp_name = re[0] if re else None
        ws_b.append([r[0], r[1], r[2], r[3], r[4], r[5], emp_name])

    ws_s = wb.create_sheet('sessions')
    ws_s.append(['id','booking_id','session_number','date','employee'])
    cur.execute('SELECT id,booking_id,session_number,date,employee_id FROM sessions WHERE booking_id IN (SELECT id FROM bookings WHERE customer_id=?) ORDER BY id', (customer_id,))
    for r in cur.fetchall():
        emp_name = None
        if r[4]:
            cur.execute('SELECT name FROM employees WHERE id=?', (r[4],))
            re = cur.fetchone()
            emp_name = re[0] if re else None
        ws_s.append([r[0], r[1], r[2], r[3], emp_name])

    ws_p = wb.create_sheet('payments')
    ws_p.append(['id','booking_id','amount','method','date','employee'])
    cur.execute('''SELECT p.id,p.booking_id,p.amount,p.method,p.date,p.employee_id FROM payments p WHERE p.booking_id IN (SELECT id FROM bookings WHERE customer_id=?) ORDER BY p.id''', (customer_id,))
    for r in cur.fetchall():
        emp_name = None
        if r[5]:
            cur.execute('SELECT name FROM employees WHERE id=?', (r[5],))
            re = cur.fetchone()
            emp_name = re[0] if re else None
        ws_p.append([r[0], r[1], r[2], r[3], r[4], emp_name])

    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f'customer_{customer_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_booking_invoice/<int:booking_id>')
@login_required
def export_booking_invoice(booking_id):
    import openpyxl
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''SELECT b.id,b.customer_id,b.package_id,b.total_sessions,b.sessions_done,b.start_date,b.employee_id,p.name,p.price,c.name,c.phone FROM bookings b JOIN packages p ON p.id=b.package_id JOIN customers c ON c.id=b.customer_id WHERE b.id=?''', (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return redirect(url_for('index'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'invoice'
    ws.append(['booking_id','customer_id','customer_name','phone','package','price','total_sessions','sessions_done','start_date','employee'])
    emp_name = None
    if row[6]:
        cur.execute('SELECT name FROM employees WHERE id=?', (row[6],))
        re = cur.fetchone()
        emp_name = re[0] if re else None
    ws.append([row[0], row[1], row[9], row[10], row[7], row[8], row[3], row[4], row[5], emp_name])
    ws2 = wb.create_sheet('sessions')
    ws2.append(['id','session_number','date','employee'])
    cur.execute('SELECT id,session_number,date,employee_id FROM sessions WHERE booking_id=? ORDER BY session_number', (booking_id,))
    for s in cur.fetchall():
        en = None
        if s[3]:
            cur.execute('SELECT name FROM employees WHERE id=?', (s[3],))
            re = cur.fetchone()
            en = re[0] if re else None
        ws2.append([s[0], s[1], s[2], en])
    ws3 = wb.create_sheet('payments')
    ws3.append(['id','amount','method','date','employee'])
    cur.execute('SELECT id,amount,method,date,employee_id FROM payments WHERE booking_id=? ORDER BY id', (booking_id,))
    for p in cur.fetchall():
        en = None
        if p[4]:
            cur.execute('SELECT name FROM employees WHERE id=?', (p[4],))
            re = cur.fetchone()
            en = re[0] if re else None
        ws3.append([p[0], p[1], p[2], p[3], en])
    ws4 = wb.create_sheet('employees')
    ws4.append(['employee','sessions_count','paid_total','cash','wallet','instapay'])
    cur.execute('SELECT e.name, COUNT(s.id) FROM sessions s LEFT JOIN employees e ON e.id=s.employee_id WHERE s.booking_id=? GROUP BY e.name', (booking_id,))
    s_map = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name', (booking_id,))
    pay_map = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, p.method, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name, p.method', (booking_id,))
    rows = cur.fetchall()
    m_map = {}
    for n, m, v in rows:
        if not n:
            continue
        d = m_map.setdefault(n, {'cash': 0, 'wallet': 0, 'instapay': 0})
        if m in d:
            d[m] = v
    names = set(list(s_map.keys()) + list(pay_map.keys()) + list(m_map.keys()))
    for n in names:
        d = m_map.get(n, {})
        ws4.append([n, s_map.get(n, 0), pay_map.get(n, 0), d.get('cash', 0), d.get('wallet', 0), d.get('instapay', 0)])
    conn.close()
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    fname = f'invoice_{booking_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    fpath = os.path.join(EXPORTS_DIR, fname)
    wb.save(fpath)
    return send_file(fpath, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/invoice/<int:booking_id>')
@login_required
def invoice(booking_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''SELECT b.id,b.customer_id,b.package_id,b.total_sessions,b.sessions_done,b.start_date,b.employee_id,p.name,p.price,c.name,c.phone,c.note FROM bookings b JOIN packages p ON p.id=b.package_id JOIN customers c ON c.id=b.customer_id WHERE b.id=?''', (booking_id,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return redirect(url_for('index'))
    cur.execute('SELECT id,session_number,date,employee_id FROM sessions WHERE booking_id=? ORDER BY session_number', (booking_id,))
    sessions = cur.fetchall()
    emp_name = None
    if row[6]:
        cur.execute('SELECT name FROM employees WHERE id=?', (row[6],))
        re = cur.fetchone()
        emp_name = re[0] if re else None
    cur.execute('SELECT e.name, COUNT(s.id) FROM sessions s LEFT JOIN employees e ON e.id=s.employee_id WHERE s.booking_id=? GROUP BY e.name', (booking_id,))
    sessions_by_emp = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name', (booking_id,))
    payments_by_emp = {r[0]: r[1] for r in cur.fetchall() if r[0]}
    cur.execute('SELECT e.name, p.method, COALESCE(SUM(p.amount),0) FROM payments p LEFT JOIN employees e ON e.id=p.employee_id WHERE p.booking_id=? GROUP BY e.name, p.method', (booking_id,))
    method_rows = cur.fetchall()
    emp_methods = {}
    for n, m, s in method_rows:
        if not n:
            continue
        d = emp_methods.setdefault(n, {'cash': 0, 'wallet': 0, 'instapay': 0})
        if m in d:
            d[m] = s
    emp_detail_list = []
    names = set(list(sessions_by_emp.keys()) + list(payments_by_emp.keys()))
    for n in names:
        ms = emp_methods.get(n, {})
        emp_detail_list.append({
            'name': n,
            'sessions_count': sessions_by_emp.get(n, 0),
            'paid_total': payments_by_emp.get(n, 0),
            'cash': ms.get('cash', 0),
            'wallet': ms.get('wallet', 0),
            'instapay': ms.get('instapay', 0),
        })
    conn.close()
    return render_template('invoice.html', booking=row, sessions=sessions, employee_name=emp_name, emp_detail_list=emp_detail_list)

@app.route("/expenses")
@login_required
def expenses():
    expenses_list = Expense.find_all_by_date_range('1970-01-01', '2100-12-31')
    total_expenses = sum(e['amount'] for e in expenses_list)
    employees = Employee.find_all_names_sorted()
    return render_template('expenses.html', 
                         expenses=[(e['id'], e['description'], e['amount'], e['category'], e['date'], e['employee_name']) for e in expenses_list], 
                         total_expenses=total_expenses,
                         employees=employees)

@app.route("/expenses/add", methods=["POST"])
@login_required
def add_expense():
    description = request.form.get('description')
    amount = request.form.get('amount')
    category = request.form.get('category')
    employee_id = request.form.get('employee_id') or session.get('employee_id')
    
    if not description or not amount or not category:
        return redirect(url_for('expenses'))
    
    try:
        amount = int(amount)
    except ValueError:
        return redirect(url_for('expenses'))
    try:
        employee_id = int(employee_id or 0)
    except Exception:
        employee_id = session.get('employee_id')
    
    today = datetime.now().strftime("%Y-%m-%d")
    Expense.create(description, amount, category, today, employee_id)
    if amount >= 500:
        Notification.create(f"مصروف عالي: {description} بمبلغ {amount} ج.م", "warning")
    return redirect(url_for('expenses'))

@app.route("/expenses/delete/<int:expense_id>")
@login_required
def delete_expense(expense_id):
    from models.base import BaseModel
    conn = BaseModel.get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('expenses'))

@app.route("/settings/whatsapp", methods=["GET", "POST"])
@manager_required
def whatsapp_settings_page():
    if request.method == "POST":
        instance_id = request.form.get("instance_id")
        api_token = request.form.get("api_token")
        sender_phone = request.form.get("sender_phone")
        is_active = 1 if request.form.get("is_active") == "on" else 0
        WhatsAppSettings.update_settings(instance_id, api_token, sender_phone, is_active)
        flash("تم حفظ إعدادات الواتساب بنجاح", "success")
    settings = WhatsAppSettings.get_settings()
    return render_template("whatsapp_settings.html", settings=settings)

if __name__ == "__main__":
    init_db()
    seed_packages()
    
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        agent_thread = threading.Thread(target=whatsapp_background_agent, daemon=True)
        agent_thread.start()

        try:
            import webbrowser
            def _open():
                time.sleep(1.5)
                try:
                    webbrowser.open('http://127.0.0.1:5007/login')
                except Exception:
                    pass
            threading.Thread(target=_open, daemon=True).start()
        except Exception:
            pass

    port = int(os.environ.get("PORT", 5007))
    debug = os.environ.get("FLASK_DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
